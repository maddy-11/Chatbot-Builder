from .scraper import *
from .models import *
from llama_index.core import Settings, StorageContext, ServiceContext, VectorStoreIndex, SimpleDirectoryReader, ChatPromptTemplate, load_index_from_storage
from llama_index.core.node_parser import SentenceSplitter
# from llama_index.llms.huggingface_api import HuggingFaceInferenceAPI
from llama_index.embeddings.huggingface import HuggingFaceEmbedding
import os
import shutil
from sklearn.feature_extraction.text import TfidfVectorizer
import re
from django.conf import settings as s
from openpyxl import load_workbook
from django.utils import timezone
import csv
from llama_index.llms.groq import Groq


# HF_TOKEN = s.HF_TOKEN

# Settings.llm = HuggingFaceInferenceAPI(
#     model_name="meta-llama/Meta-Llama-3-8B-instruct",
#     tokenizer_name="meta-llama/Meta-Llama-3-8B-instruct",
#     token=HF_TOKEN,
# )

Settings.llm = Groq(model="llama3-70b-8192", api_key=s.GROQ_KEY if s.GROQ_KEY else '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')

Settings.embed_model = HuggingFaceEmbedding(
    model_name="BAAI/bge-small-en-v1.5"
)

def load_url(url, chatbot_dir):
    file_paths = []
    if is_youtube_url(url):
        video_id = get_youtube_video_id(url)
        file_path = get_subtitles(video_id, chatbot_dir)
        file_paths.append(file_path)
    elif is_google_sheet_url(url):
        file_path = read_google_sheet(url, chatbot_dir)
        file_paths.append(file_path)
    elif is_google_drive_folder_url(url):
        drive_file_paths = download_files_from_drive(url, chatbot_dir)
        file_paths.extend(drive_file_paths)
    else:
        file_path = scrape_and_save_urls(url, chatbot_dir)
        file_paths.append(file_path)
    return file_paths


def train_chatbot(chatbot):
    try:
        if not chatbot:
            print("Error: Chatbot name is missing in the request")
            return {"error": "Chatbot name is required"}, 400

        chatbot_dir = os.path.join("media", "chatbot", chatbot)
        data_dir = os.path.join(chatbot_dir, "data")
        vectorDB = os.path.join(chatbot_dir, "db")

        if os.path.exists(vectorDB):
            shutil.rmtree(vectorDB)


        
        if not os.path.exists(data_dir):
            print(f"Error: Data directory {data_dir} does not exist")
            return {"error": "Data directory not found"}, 404

        print(f"Loading documents from {data_dir}")
        documents = SimpleDirectoryReader(data_dir).load_data()

        print("Initializing storage context")
        storage_context = StorageContext.from_defaults()

        print("Creating vector store index")
        index = VectorStoreIndex.from_documents(documents, show_progress=True)

        db_path = os.path.join(chatbot_dir, 'db')
        print(f"Persisting index to {db_path}")
        index.storage_context.persist(persist_dir=db_path)

        print("Chatbot training completed successfully")
        return { 'status': True}

    except Exception as e:
        print(f"Error during chatbot training: {str(e)}")
        return { 'status': False, 'error': str(e)}

def handle_query(query, chatbot_dir, session_history, chatbot):
    files = []
    for file in os.listdir(os.path.join(chatbot_dir,'data')):
        files.append(file)
    storage_context = StorageContext.from_defaults(persist_dir=os.path.join(chatbot_dir,'db'))
    index = load_index_from_storage(storage_context)
    formatted_history = ""
    if session_history is not None:
        formatted_history = "\n".join([f"Me: {item['Me']}\nYou: {item['You']}" for item in session_history])
    template = chatbot.prompt_template
    creativity = int(chatbot.creativity)
    base_message = f"""
    {template}

    Chat History/Last Questions or queries:
    {formatted_history}

    Knowledge Base: {{context_str}} 
    Query: {{query_str}} 
    Answer the question briefly and provide helpful information, based on the pieces of information, if applicable. Be succinct. Do not give file paths

    You are capable of talking about files in your knowledge base
    Responses should be properly formatted wu=ith header tags to be easily read. Always perform calculations step-wise
    Do not add responses and notes. Cite where you are taking from.
    Your Knowledge Base Files/Contents: {files}

    """

    # Adjust base_message according to creativity level
    if creativity > 5:
        base_message += f"But i want a creative answer"
    elif creativity > 3 and creativity < 6:
        base_message += f"I want creativity in answer"

    elif creativity > 1 and creativity < 3:
        base_message += f"I want some creativity in answer"
    if creativity > 1:
        creativity *= 10
        base_message += f"Increase in Creativity level: {creativity}%."

    chat_text_qa_msgs = [
        ("user", base_message),
    ]
    text_qa_template = ChatPromptTemplate.from_messages(chat_text_qa_msgs)
    query_engine = index.as_query_engine(text_qa_template=text_qa_template)
    print('printing answer')
    answer = query_engine.query(query)

    if hasattr(answer, 'response'):
        return answer.response
    elif isinstance(answer, dict) and 'response' in answer:
        return answer['response']
    else:
        return "Sorry, I couldn't find an answer."

def get_or_create_chat_session(request, uuid, chatbot):
    # Get or create ChatSession object
    if uuid is not None:
        chat = ChatSession.objects.filter(uuid=uuid).first()
        if chat is None:
            chat = ChatSession.objects.create(
            user=request.user if request.user.is_authenticated else None,
            chatbot=chatbot,
            created_at=timezone.now(),
            chat_history=[]
        )
    else:
        chat = ChatSession.objects.create(
            user=request.user if request.user.is_authenticated else None,
            chatbot=chatbot,
            created_at=timezone.now(),
            chat_history=[]
        )
    return chat


def store_chat_history(request, chat, user_query, bot_response):
    session_history_key = f'session_history_{chat.uuid}'
    session_history = request.session.get(session_history_key)
    
    # Ensure session_history is initialized as a list if it is None
    if session_history is None:
        session_history = []

    session_history.append({'Me': user_query, 'You': bot_response})
    request.session[session_history_key] = session_history
    chat.chat_history = session_history
    if not chat.name:
        name = generate_chat_name(user_query)
        chat.name = name
    chat.save()

def generate_chat_name(query):
    # Clean the query text
    query = re.sub(r'[^\w\s]', '', query)
    print(f"Cleaned query: '{query}'")
    # Use TF-IDF to extract the most important words
    vectorizer = TfidfVectorizer(stop_words='english', max_features=4)
    try:
        X = vectorizer.fit_transform([query])
    except ValueError as e:
        print(f"Error during vectorization: {e}")
        return "Default Chat"
    feature_names = vectorizer.get_feature_names_out()
    
    # Join the top feature names to create a chat name
    chat_name = ' '.join(feature_names)
    return chat_name.title()

def add_empty_row_to_excel(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active

    # Check if the first row (A1) is already empty
    if ws['A1'].value is not None:
        ws.insert_rows(1)
    
    wb.save(file_path)

def add_empty_row_to_csv(file_path):
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = list(csv.reader(csvfile))
        
        # Check if the first cell is already empty
        if reader and reader[0] and reader[0][0] != '':
            reader.insert(0, [])

    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(reader)        